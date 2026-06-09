import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import spreadsheet  # noqa: E402


class SpreadsheetAppendTests(unittest.TestCase):
    def _write_fixture(self, rows: list[tuple[str, str, str]]) -> Path:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        path = Path(tmp.name)
        tmp.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for col, row, value in rows:
            ws[f"{col}{row}"] = value
        wb.save(path)
        wb.close()
        return path

    def test_skips_partially_filled_row_and_appends_after_full_block(self):
        path = self._write_fixture([
            ("A", 1, "1"),
            ("B", 1, "2"),
            ("C", 1, "3"),
            ("A", 2, "4"),
        ])
        try:
            spreadsheet.append_rows(
                path,
                "Sheet1",
                [{"A": "new-a", "B": "new-b", "C": "new-c"}],
            )
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Sheet1"]
            self.assertEqual(ws["A3"].value, "new-a")
            self.assertEqual(ws["B3"].value, "new-b")
            self.assertEqual(ws["C3"].value, "new-c")
            wb.close()
        finally:
            path.unlink(missing_ok=True)

    def test_ignores_data_outside_write_columns(self):
        path = self._write_fixture([("D", 1, "outside")])
        try:
            spreadsheet.append_rows(
                path,
                "Sheet1",
                [{"A": "a1", "B": "b1"}],
            )
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Sheet1"]
            self.assertEqual(ws["A1"].value, "a1")
            self.assertEqual(ws["B1"].value, "b1")
            wb.close()
        finally:
            path.unlink(missing_ok=True)

    def test_appends_multiple_rows_sequentially(self):
        path = self._write_fixture([("A", 1, "existing")])
        try:
            spreadsheet.append_rows(
                path,
                "Sheet1",
                [
                    {"A": "row2", "B": "b2"},
                    {"A": "row3", "B": "b3"},
                ],
            )
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb["Sheet1"]
            self.assertEqual(ws["A2"].value, "row2")
            self.assertEqual(ws["B2"].value, "b2")
            self.assertEqual(ws["A3"].value, "row3")
            self.assertEqual(ws["B3"].value, "b3")
            wb.close()
        finally:
            path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
