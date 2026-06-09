import threading
import re
from pathlib import Path

import openpyxl

# Per-document write locks to prevent concurrent corruption
_file_locks: dict[str, threading.Lock] = {}
_locks_mutex = threading.Lock()


def _get_lock(doc_id: str) -> threading.Lock:
    with _locks_mutex:
        if doc_id not in _file_locks:
            _file_locks[doc_id] = threading.Lock()
        return _file_locks[doc_id]


def _doc_id_from_path(path: Path) -> str:
    return path.stem


def _load_or_create(path: Path) -> openpyxl.Workbook:
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _ensure_sheet(wb: openpyxl.Workbook, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


def _sync_headers(ws, new_keys: list[str]) -> list[str]:
    """Return the full ordered header list, appending any keys not yet present."""
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        for i, key in enumerate(new_keys, 1):
            ws.cell(1, i).value = key
        return list(new_keys)

    existing = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    headers = [h for h in existing if h is not None]

    for key in new_keys:
        if key not in headers:
            headers.append(key)
            ws.cell(1, len(headers)).value = key

    return headers


_COLUMN_RE = re.compile(r"^[A-Za-z]+$")


def _is_column_name(value: str) -> bool:
    return bool(_COLUMN_RE.match((value or "").strip()))


def _cell_is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _last_occupied_row_in_column(ws, column_letter: str) -> int:
    """Highest row index with a non-empty value in the given column, or 0."""
    column = (column_letter or "").upper().strip()
    if not column:
        return 0
    for row in range(ws.max_row, 0, -1):
        if not _cell_is_empty(ws[f"{column}{row}"].value):
            return row
    return 0


def _first_free_row_for_columns(ws, column_letters: list[str]) -> int:
    """
    First row where every target column is empty.

    Matches Google Sheets behaviour: take the deepest occupied row across the
    write columns and append on the next row, ignoring data outside the range.
    """
    columns = {(col or "").upper().strip() for col in column_letters if (col or "").strip()}
    if not columns:
        return 1
    max_occupied = max(_last_occupied_row_in_column(ws, col) for col in columns)
    return max_occupied + 1


def append_rows(path: Path, sheet_name: str, field_rows: list[dict[str, str]]) -> None:
    doc_id = _doc_id_from_path(path)
    with _get_lock(doc_id):
        wb = _load_or_create(path)
        ws = _ensure_sheet(wb, sheet_name)

        # If all keys are spreadsheet columns (A, B, ... AA), write directly
        # to those columns without creating an implicit header row.
        all_keys_flat = [k for row in field_rows for k in row]
        if all_keys_flat and all(_is_column_name(k) for k in all_keys_flat):
            write_columns = list(dict.fromkeys(k.upper().strip() for k in all_keys_flat))
            target_row = _first_free_row_for_columns(ws, write_columns)
            for row_data in field_rows:
                for col, value in row_data.items():
                    ws[f"{col.upper().strip()}{target_row}"] = value
                target_row += 1
            wb.save(path)
            return

        all_keys = list(dict.fromkeys(k for row in field_rows for k in row))
        headers = _sync_headers(ws, all_keys)

        for row_data in field_rows:
            ws.append([row_data.get(h, "") for h in headers])

        wb.save(path)


def get_rows(path: Path, sheet_name: str) -> list[dict]:
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [
        {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)}
        for row in rows[1:]
    ]


def list_sheets(path: Path) -> list[str]:
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    finally:
        wb.close()


def init_sheet(path: Path, sheet_name: str, columns: list[str]) -> None:
    doc_id = _doc_id_from_path(path)
    with _get_lock(doc_id):
        wb = _load_or_create(path)
        ws = _ensure_sheet(wb, sheet_name)
        # Write header only if sheet is empty
        if ws.max_row == 0 or ws.cell(1, 1).value is None:
            for i, col in enumerate(columns, 1):
                ws.cell(1, i).value = col
        wb.save(path)
